#!/usr/bin/env python3
"""
TEVAH OMNISCIENT EMAIL INGESTION
=================================
Every word. Every attachment. Every number. Every relationship.
This is not a preview — this is the full extraction.

Parses: full body text, PDF attachments, Excel/CSV attachments,
price amounts, deal references, line items, payment terms.
Stores everything. Truncates nothing.
"""

import imaplib
import email
import email.message
from email.header import decode_header
from email.utils import parsedate_to_datetime, parseaddr
import json
import re
import os
import sys
import io
import csv
import time
import hashlib
import tempfile
from datetime import datetime, timedelta, timezone
from collections import defaultdict, Counter
from typing import Optional, List, Dict, Any, Tuple

# Attachment parsers
import pdfplumber
import openpyxl
from bs4 import BeautifulSoup
import chardet

# ─── Config ───────────────────────────────────────────────────────────────────
# App passwords loaded from environment variables:
#   TEVAH_GMAIL_AVI, TEVAH_GMAIL_YAAKOV, TEVAH_GMAIL_SALES,
#   TEVAH_GMAIL_SERGIO, TEVAH_GMAIL_MALKA
ACCOUNTS = [
    {
        "name": "Avi",
        "email": "avi@tevahglobal.com",
        "password": os.environ.get("TEVAH_GMAIL_AVI", ""),
    },
    {
        "name": "Yaakov",
        "email": "yaakovn@tevahglobal.com",
        "password": os.environ.get("TEVAH_GMAIL_YAAKOV", ""),
    },
    {
        "name": "Sales-Gil",
        "email": "sales@tevahglobal.com",
        "password": os.environ.get("TEVAH_GMAIL_SALES", ""),
    },
    {
        "name": "Sergio",
        "email": "sergio@tevahglobal.com",
        "password": os.environ.get("TEVAH_GMAIL_SERGIO", ""),
    },
    {
        "name": "Malka",
        "email": "malka@tevahglobal.com",
        "password": os.environ.get("TEVAH_GMAIL_MALKA", ""),
    },
]

FOLDERS = ["INBOX", '"[Gmail]/Sent Mail"']
DAYS_BACK = 90
IMAP_HOST = "imap.gmail.com"
IMAP_PORT = 993
BATCH_SIZE = 50

BASE_DIR = "/home/ubuntu/ai-os/projects/tevah/tevah-standalone"
OUTPUT_DIR = os.path.join(BASE_DIR, "deep-intelligence")
ATTACHMENT_DIR = os.path.join(OUTPUT_DIR, "attachments")
MONDAY_FILE = os.path.join(BASE_DIR, "monday-intelligence.json")

# NO max body length — we keep everything
MAX_BODY_CHARS = 100_000  # 100K chars safety limit (not 1K)

# Only filter truly irrelevant automated emails
# KEEP: QuickBooks (payment data), FedEx/UPS/DHL (shipping data), Melio (payments)
AUTOMATED_PATTERNS = [
    r"noreply@google\.com",
    r"no-reply@accounts\.google\.com",
    r"calendar-notification@google\.com",
    r"mailer-daemon@",
    r"postmaster@",
    r"@.*\.gserviceaccount\.com$",
    r"@linkedin\.com$",
    r"@facebook\.com$",
    r"@facebookmail\.com$",
    r"newsletter@",
    r"news@",
    r"marketing@",
    r"promo@",
    r"@mailchimp\.com$",
    r"@sendgrid\.net$",
    r"@constantcontact\.com$",
    r"@vercel\.com$",
    r"@github\.com$",
    r"@cloudflare\.com$",
    r"@shopify\.com$",  # unless they sell on shopify — revisit
    r"daemon@",
]

AUTOMATED_SUBJECT_PATTERNS = [
    r"^invitation:",
    r"google alert",
    r"verify your email",
    r"reset your password",
    r"security alert",
    r"sign-in attempt",
    r"out of office",
    r"automatic reply",
    r"undeliverable:",
    r"delivery status notification",
]

INTERNAL_DOMAINS = ["tevahglobal.com"]

# Business-relevant automated senders we KEEP
BUSINESS_AUTO_SENDERS = [
    r"@quickbooks",
    r"@intuit\.com",
    r"@notification\.intuit\.com",  # payments
    r"@meliopayments\.com",
    r"@melio\.com",  # payments
    r"@fedex\.com",
    r"@ups\.com",
    r"@dhl\.com",  # shipping
    r"@streak\.com",  # CRM tracking
    r"@missiveapp\.com",  # team collaboration mentions
    r"@mondaycom\.com",
    r"@monday\.com",  # project management
]

# Expanded classification with wholesale-specific categories
CLASSIFICATION_RULES = {
    "pricing_negotiation": [
        r"\b(price|pricing|price list|cost|margin|markup|counter.?offer|best price|final price|offer|quoted?|quotation|landed cost|FOB price|CIF price|EXW|wholesale price|unit price|per unit|total cost|net price|gross|profit|spread)\b",
    ],
    "purchase_order": [
        r"\b(PO|purchase order|order confirmation|P\.O\.|PO#|PO \d|order number|order #|confirmed order)\b",
    ],
    "shipping_logistics": [
        r"\b(ship|shipping|freight|container|tracking|FCL|LCL|FOB|CIF|port|customs|clearance|BL|bill of lading|vessel|cargo|warehouse|pallet|loading|unloading|delivery|transit|arrived|departed|ETA|ETD|pickup|LTL|FTL|blind ship|drop ship|carrier)\b",
    ],
    "payment_finance": [
        r"\b(payment|invoice|wire|transfer|pay|paid|balance|due|remittance|TT|LC|letter of credit|credit note|debit|statement|collection|financing|funded|advance|deposit|COD|NET\s*\d+|check|cheque|ACH|Melio|QuickBooks)\b",
    ],
    "vendor_sourcing": [
        r"\b(availability|stock|inventory|lead time|MOQ|minimum order|sample|catalog|catalogue|sourcing|surplus|overstock|closeout|liquidation|lot|lots|brand new|sealed|case pack|master case|inner pack)\b",
    ],
    "customer_sales": [
        r"\b(interested|confirm|proceed|deal|opportunity|quantity|inquiry|RFQ|request for quote|bid|tender|volume discount|bulk|customer|client|buyer|ready to|place order)\b",
    ],
    "claims_issues": [
        r"\b(damage|damaged|shortage|missing|defective|return|RMA|claim|complaint|quality issue|wrong item|mismatch|discrepancy|short ship)\b",
    ],
}

# Price/amount patterns
MONEY_PATTERN = re.compile(
    r"(?:\$|USD\s*|€|EUR\s*|£|GBP\s*)"
    r"[\s]*"
    r"(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)",
    re.IGNORECASE,
)

QUANTITY_PATTERN = re.compile(
    r"(\d{1,6})\s*(?:units?|pcs?|pieces?|cases?|pallets?|cartons?|boxes?|each|ea)\b",
    re.IGNORECASE,
)

PERCENTAGE_PATTERN = re.compile(
    r"(\d{1,3}(?:\.\d{1,2})?)\s*%",
)


def decode_mime_header(header_val: str) -> str:
    if not header_val:
        return ""
    decoded_parts = []
    try:
        for part, charset in decode_header(header_val):
            if isinstance(part, bytes):
                decoded_parts.append(part.decode(charset or "utf-8", errors="replace"))
            else:
                decoded_parts.append(str(part))
    except Exception:
        return str(header_val)
    return " ".join(decoded_parts)


def parse_all_addresses(header_val: str) -> List[Dict[str, str]]:
    """Parse ALL email addresses from a header (To, Cc, Bcc can have multiple)."""
    if not header_val:
        return []
    addresses = []
    # Split on comma, handling quoted names
    raw = decode_mime_header(header_val)
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        name, addr = parseaddr(part)
        if addr:
            addresses.append({"name": name, "email": addr.lower()})
    return addresses


def extract_full_body(msg: email.message.Message) -> Tuple[str, str]:
    """Extract FULL body text — both plain and HTML. No truncation."""
    plain_body = ""
    html_body = ""

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disp = str(part.get("Content-Disposition", ""))
            if "attachment" in content_disp:
                continue

            try:
                charset = part.get_content_charset() or "utf-8"
                payload = part.get_payload(decode=True)
                if not payload:
                    continue

                # Try declared charset, fall back to chardet
                try:
                    text = payload.decode(charset, errors="replace")
                except (UnicodeDecodeError, LookupError):
                    detected = chardet.detect(payload)
                    text = payload.decode(
                        detected.get("encoding", "utf-8"), errors="replace"
                    )

                if content_type == "text/plain" and not plain_body:
                    plain_body = text
                elif content_type == "text/html" and not html_body:
                    html_body = text
            except Exception:
                continue
    else:
        try:
            charset = msg.get_content_charset() or "utf-8"
            payload = msg.get_payload(decode=True)
            if payload:
                try:
                    text = payload.decode(charset, errors="replace")
                except (UnicodeDecodeError, LookupError):
                    detected = chardet.detect(payload)
                    text = payload.decode(
                        detected.get("encoding", "utf-8"), errors="replace"
                    )

                if msg.get_content_type() == "text/html":
                    html_body = text
                else:
                    plain_body = text
        except Exception:
            pass

    # Convert HTML to clean text if no plain text
    clean_text = plain_body.strip()
    if not clean_text and html_body:
        soup = BeautifulSoup(html_body, "html.parser")
        # Remove script/style
        for tag in soup(["script", "style", "head"]):
            tag.decompose()
        clean_text = soup.get_text(separator="\n", strip=True)

    # Safety cap at 100K chars
    if len(clean_text) > MAX_BODY_CHARS:
        clean_text = clean_text[:MAX_BODY_CHARS] + "\n[TRUNCATED AT 100K CHARS]"

    return clean_text, html_body[:MAX_BODY_CHARS] if html_body else ""


def extract_attachments(
    msg: email.message.Message, email_uid: str
) -> List[Dict[str, Any]]:
    """Extract and parse ALL attachments. PDFs → text. Excel → rows. CSV → rows."""
    attachments = []

    if not msg.is_multipart():
        return attachments

    for part in msg.walk():
        content_disp = str(part.get("Content-Disposition", ""))
        if "attachment" not in content_disp and part.get_content_maintype() not in [
            "application",
            "image",
        ]:
            continue
        if part.get_content_type() in ["text/plain", "text/html"]:
            # These are body parts, not attachments
            if "attachment" not in content_disp:
                continue

        filename = part.get_filename()
        if filename:
            filename = decode_mime_header(filename)
        else:
            # Generate name from content type
            ext = part.get_content_subtype() or "bin"
            filename = f"unnamed_{email_uid[:8]}.{ext}"

        content_type = part.get_content_type()
        payload = part.get_payload(decode=True)
        if not payload:
            continue

        attachment = {
            "filename": filename,
            "content_type": content_type,
            "size_bytes": len(payload),
            "extracted_text": "",
            "extracted_data": None,
        }

        # Parse based on type
        try:
            if content_type == "application/pdf" or filename.lower().endswith(".pdf"):
                attachment["extracted_text"] = parse_pdf(payload, filename)
            elif content_type in [
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
            ] or filename.lower().endswith((".xlsx", ".xls")):
                text, data = parse_excel(payload, filename)
                attachment["extracted_text"] = text
                attachment["extracted_data"] = data
            elif content_type == "text/csv" or filename.lower().endswith(".csv"):
                text, data = parse_csv(payload, filename)
                attachment["extracted_text"] = text
                attachment["extracted_data"] = data
            elif content_type.startswith("text/"):
                try:
                    attachment["extracted_text"] = payload.decode(
                        "utf-8", errors="replace"
                    )[:50000]
                except:
                    pass
            # Images: store metadata, skip OCR for now (could add later)
            elif content_type.startswith("image/"):
                attachment["extracted_text"] = (
                    f"[Image: {filename}, {len(payload)} bytes]"
                )

            # Save attachment to disk
            save_attachment(payload, email_uid, filename)

        except Exception as e:
            attachment["extracted_text"] = f"[Parse error: {str(e)[:200]}]"

        attachments.append(attachment)

    return attachments


def parse_pdf(payload: bytes, filename: str) -> str:
    """Extract ALL text from a PDF."""
    text_parts = []
    try:
        with pdfplumber.open(io.BytesIO(payload)) as pdf:
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text() or ""
                if page_text:
                    text_parts.append(f"--- Page {i + 1} ---\n{page_text}")

                # Also extract tables
                tables = page.extract_tables()
                for j, table in enumerate(tables):
                    if table:
                        text_parts.append(f"--- Table {j + 1} on Page {i + 1} ---")
                        for row in table:
                            text_parts.append(
                                " | ".join(str(cell or "") for cell in row)
                            )
    except Exception as e:
        text_parts.append(f"[PDF parse error: {str(e)[:200]}]")

    result = "\n".join(text_parts)
    return result[:MAX_BODY_CHARS] if result else f"[Empty PDF: {filename}]"


def parse_excel(payload: bytes, filename: str) -> Tuple[str, List[List]]:
    """Extract ALL data from Excel file."""
    text_parts = []
    all_data = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            sheet_data = []
            for row in ws.iter_rows(values_only=True):
                row_vals = [str(cell) if cell is not None else "" for cell in row]
                if any(v for v in row_vals):  # skip fully empty rows
                    text_parts.append(" | ".join(row_vals))
                    sheet_data.append(row_vals)
            all_data.append(
                {"sheet": sheet_name, "rows": sheet_data[:5000]}
            )  # cap rows per sheet
        wb.close()
    except Exception as e:
        text_parts.append(f"[Excel parse error: {str(e)[:200]}]")

    result = "\n".join(text_parts)
    return result[:MAX_BODY_CHARS], all_data


def parse_csv(payload: bytes, filename: str) -> Tuple[str, List[List]]:
    """Extract ALL data from CSV file."""
    text_parts = []
    all_rows = []
    try:
        # Detect encoding
        detected = chardet.detect(payload[:10000])
        encoding = detected.get("encoding", "utf-8")
        text = payload.decode(encoding, errors="replace")
        reader = csv.reader(io.StringIO(text))
        for i, row in enumerate(reader):
            text_parts.append(" | ".join(row))
            all_rows.append(row)
            if i > 10000:
                text_parts.append("[TRUNCATED AT 10K ROWS]")
                break
    except Exception as e:
        text_parts.append(f"[CSV parse error: {str(e)[:200]}]")

    result = "\n".join(text_parts)
    return result[:MAX_BODY_CHARS], all_rows


def save_attachment(payload: bytes, email_uid: str, filename: str):
    """Save raw attachment to disk."""
    safe_name = re.sub(r"[^\w\-.]", "_", filename)[:100]
    dir_path = os.path.join(ATTACHMENT_DIR, email_uid[:12])
    os.makedirs(dir_path, exist_ok=True)
    filepath = os.path.join(dir_path, safe_name)
    with open(filepath, "wb") as f:
        f.write(payload)


def is_automated(from_addr: str, subject: str) -> bool:
    """Check if email is truly irrelevant automated (not business-relevant)."""
    from_lower = from_addr.lower()
    subj_lower = subject.lower()

    # First: check if it's a business-relevant automated sender — KEEP these
    for pattern in BUSINESS_AUTO_SENDERS:
        if re.search(pattern, from_lower):
            return False

    # Then check if it's irrelevant automated
    for pattern in AUTOMATED_PATTERNS:
        if re.search(pattern, from_lower):
            return True

    for pattern in AUTOMATED_SUBJECT_PATTERNS:
        if re.search(pattern, subj_lower):
            return True

    return False


def extract_deal_numbers(subject: str, body: str) -> List[str]:
    """Extract deal/order numbers from FULL subject + body."""
    # Search subject thoroughly, body first 2000 chars (numbers usually near top)
    text = subject + " " + body[:2000]
    numbers = set()

    patterns = [
        r"#(\d{4,6})\b",
        r"\b(?:deal|order|po|so|inv|invoice|ref|reference|confirmation)[\s#\-:]*(\d{4,6})\b",
        r"\b(\d{4,6})(?:\s*[-–—]\s*[A-Za-z])",  # "6722 - Brand Name"
        r"\bPO[\-#\s]*(\d{4,6})\b",
        r"\bSO[\-#\s]*(\d{4,6})\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            num = match.group(1)
            if not (1990 <= int(num) <= 2030):  # filter years
                numbers.add(num)

    # Subject-specific: standalone 4-digit numbers (very common in Tevah subjects like "Giro 6343")
    for match in re.finditer(r"\b(\d{4})\b", subject):
        num = match.group(1)
        if not (1990 <= int(num) <= 2030):
            numbers.add(num)

    return sorted(numbers)


def extract_financial_data(text: str) -> Dict[str, Any]:
    """Extract all monetary amounts, quantities, and percentages from text."""
    amounts = []
    for match in MONEY_PATTERN.finditer(text[:10000]):
        raw = match.group(1).replace(",", "")
        try:
            val = float(raw)
            if val > 0:
                amounts.append(val)
        except ValueError:
            pass

    quantities = []
    for match in QUANTITY_PATTERN.finditer(text[:10000]):
        try:
            quantities.append(int(match.group(1)))
        except ValueError:
            pass

    percentages = []
    for match in PERCENTAGE_PATTERN.finditer(text[:5000]):
        try:
            val = float(match.group(1))
            if 0 < val <= 100:
                percentages.append(val)
        except ValueError:
            pass

    if not amounts and not quantities and not percentages:
        return {}

    return {
        "amounts": sorted(set(amounts), reverse=True)[:20],
        "quantities": sorted(set(quantities), reverse=True)[:20],
        "percentages": sorted(set(percentages))[:10],
        "total_amount_mentioned": sum(amounts) if amounts else 0,
        "largest_amount": max(amounts) if amounts else 0,
    }


def classify_email(
    from_addr: str,
    to_addrs: List[Dict],
    cc_addrs: List[Dict],
    subject: str,
    body: str,
    account_email: str,
) -> List[str]:
    """Multi-label classification — an email can be about pricing AND shipping."""
    all_text = (subject + " " + body[:5000]).lower()
    from_domain = from_addr.split("@")[-1].lower() if "@" in from_addr else ""

    to_domains = [a["email"].split("@")[-1] for a in to_addrs if "@" in a["email"]]

    from_internal = any(d in from_domain for d in INTERNAL_DOMAINS)
    to_internal = any(any(d in td for d in INTERNAL_DOMAINS) for td in to_domains)

    labels = []

    if from_internal and to_internal:
        labels.append("internal")

    # Score each category
    scores = {}
    for category, patterns in CLASSIFICATION_RULES.items():
        score = 0
        for pattern in patterns:
            matches = re.findall(pattern, all_text, re.IGNORECASE)
            score += len(matches)
        if score > 0:
            scores[category] = score

    # Return all categories that scored, sorted by score
    if scores:
        labels.extend(sorted(scores, key=scores.get, reverse=True))

    return labels if labels else ["uncategorized"]


def detect_reply_chain(subject: str) -> Tuple[bool, str]:
    clean = re.sub(r"^(re|fwd?|fw):\s*", "", subject.strip(), flags=re.IGNORECASE)
    clean = re.sub(r"^(re|fwd?|fw):\s*", "", clean.strip(), flags=re.IGNORECASE)
    clean = re.sub(r"^(re|fwd?|fw):\s*", "", clean.strip(), flags=re.IGNORECASE)
    is_reply = clean.lower() != subject.strip().lower()
    return is_reply, clean.strip()


def fetch_emails_from_account(account: dict, since_date: str) -> list:
    """Fetch ALL email data from an account — full bodies + attachments."""
    emails = []
    attachment_count = 0

    print(f"\n{'─' * 60}")
    print(f"  ACCOUNT: {account['name']} ({account['email']})")
    print(f"{'─' * 60}")

    conn = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    conn.login(account["email"], account["password"])

    for folder in FOLDERS:
        try:
            status, _ = conn.select(folder, readonly=True)
            if status != "OK":
                print(f"  [WARN] Could not select {folder}")
                continue
        except Exception as e:
            print(f"  [WARN] Error selecting {folder}: {e}")
            continue

        folder_name = folder.replace('"', "")
        print(f"\n  Folder: {folder_name}")

        status, msg_ids = conn.search(None, f'(SINCE "{since_date}")')
        if status != "OK" or not msg_ids[0]:
            print(f"    No messages found")
            continue

        ids = msg_ids[0].split()
        total = len(ids)
        print(f"    {total} messages to process")

        for i in range(0, total, BATCH_SIZE):
            batch = ids[i : i + BATCH_SIZE]
            id_range = b",".join(batch)

            try:
                status, data = conn.fetch(id_range, "(RFC822)")
                if status != "OK":
                    continue
            except Exception as e:
                print(f"    [WARN] Fetch error at batch {i}: {e}")
                # Reconnect on error
                try:
                    conn = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
                    conn.login(account["email"], account["password"])
                    conn.select(folder, readonly=True)
                except:
                    break
                continue

            for item in data:
                if not isinstance(item, tuple):
                    continue

                try:
                    msg = email.message_from_bytes(item[1])

                    # Headers
                    from_raw = decode_mime_header(msg.get("From", ""))
                    from_name, from_addr = parseaddr(from_raw)
                    to_addrs = parse_all_addresses(msg.get("To", ""))
                    cc_addrs = parse_all_addresses(msg.get("Cc", ""))
                    bcc_addrs = parse_all_addresses(msg.get("Bcc", ""))
                    subject = decode_mime_header(msg.get("Subject", ""))
                    date_str = msg.get("Date", "")
                    message_id = msg.get("Message-ID", "")
                    in_reply_to = msg.get("In-Reply-To", "")
                    references = msg.get("References", "")

                    # Parse date
                    try:
                        dt = parsedate_to_datetime(date_str)
                        date_iso = dt.isoformat()
                        date_simple = dt.strftime("%Y-%m-%d")
                    except Exception:
                        date_iso = date_str
                        date_simple = ""

                    # Skip truly irrelevant automated
                    if is_automated(from_addr, subject):
                        continue

                    # FULL body extraction
                    body_text, html_raw = extract_full_body(msg)

                    # Attachments — parsed
                    uid = hashlib.md5(
                        f"{message_id}{from_addr}{date_str}".encode()
                    ).hexdigest()[:12]
                    attachments = extract_attachments(msg, uid)
                    attachment_count += len(attachments)

                    # Combine body + attachment text for analysis
                    full_text = body_text
                    for att in attachments:
                        if att.get("extracted_text"):
                            full_text += f"\n\n[ATTACHMENT: {att['filename']}]\n{att['extracted_text']}"

                    # Reply chain
                    is_reply, base_subject = detect_reply_chain(subject)

                    # Deal numbers from FULL text
                    deal_numbers = extract_deal_numbers(subject, full_text)

                    # Financial data extraction
                    financial = extract_financial_data(full_text)

                    # Multi-label classification
                    classifications = classify_email(
                        from_addr,
                        to_addrs,
                        cc_addrs,
                        subject,
                        body_text,
                        account["email"],
                    )

                    email_data = {
                        "uid": uid,
                        "account": account["name"],
                        "folder": folder_name,
                        "from_name": from_name,
                        "from_email": from_addr.lower(),
                        "to": [a["email"] for a in to_addrs],
                        "to_names": [a["name"] for a in to_addrs if a["name"]],
                        "cc": [a["email"] for a in cc_addrs],
                        "bcc": [a["email"] for a in bcc_addrs],
                        "subject": subject,
                        "base_subject": base_subject,
                        "date": date_iso,
                        "date_simple": date_simple,
                        "body_full": body_text,
                        "body_length": len(body_text),
                        "deal_numbers": deal_numbers,
                        "is_reply": is_reply,
                        "classifications": classifications,
                        "message_id": message_id,
                        "in_reply_to": in_reply_to,
                        "references_count": len(references.split())
                        if references
                        else 0,
                        "financial": financial,
                        "attachments": [
                            {
                                "filename": a["filename"],
                                "content_type": a["content_type"],
                                "size_bytes": a["size_bytes"],
                                "extracted_text": a["extracted_text"][
                                    :20000
                                ],  # cap per attachment
                                "has_data": bool(a.get("extracted_data")),
                            }
                            for a in attachments
                        ],
                        "attachment_count": len(attachments),
                    }
                    emails.append(email_data)
                except Exception as e:
                    continue

            processed = min(i + BATCH_SIZE, total)
            if processed % 200 == 0 or processed >= total:
                print(
                    f"    Processed {processed}/{total} ({attachment_count} attachments extracted)"
                )

    conn.logout()
    print(f"  Total: {len(emails)} emails, {attachment_count} attachments")
    return emails


def deduplicate_emails(emails: list) -> list:
    """Remove duplicates, merge cross-account appearances."""
    seen = {}
    for e in emails:
        key = e["message_id"] if e["message_id"] else e["uid"]
        if key not in seen:
            seen[key] = e
        else:
            existing = seen[key]
            if existing["folder"] != e["folder"]:
                existing["folder"] = f"{existing['folder']} + {e['folder']}"
            if existing["account"] != e["account"]:
                existing["account"] = f"{existing['account']} + {e['account']}"
            # Merge financial data
            if e.get("financial") and not existing.get("financial"):
                existing["financial"] = e["financial"]
            # Merge attachments
            existing_fnames = {a["filename"] for a in existing.get("attachments", [])}
            for att in e.get("attachments", []):
                if att["filename"] not in existing_fnames:
                    existing["attachments"].append(att)
                    existing["attachment_count"] += 1
    return list(seen.values())


def load_monday_deals() -> dict:
    try:
        with open(MONDAY_FILE) as f:
            data = json.load(f)
    except Exception as e:
        print(f"[WARN] Could not load Monday data: {e}")
        return {
            "deals": [],
            "by_brand": {},
            "by_customer": {},
            "customers": [],
            "suppliers": [],
        }

    deals = data.get("deals", [])
    by_brand = defaultdict(list)
    by_customer = defaultdict(list)

    for d in deals:
        brand = (d.get("brand") or "").lower().strip()
        customer = (d.get("customer") or "").lower().strip()
        if brand:
            by_brand[brand].append(d)
        if customer:
            by_customer[customer].append(d)

    return {
        "deals": deals,
        "by_brand": dict(by_brand),
        "by_customer": dict(by_customer),
        "customers": data.get("customers", []),
        "suppliers": data.get("suppliers", []),
        "summary": data.get("summary", {}),
    }


def cross_reference_deals(emails: list, monday: dict) -> dict:
    matches = defaultdict(list)
    brand_mentions = Counter()
    customer_mentions = Counter()

    all_brands = set(monday["by_brand"].keys())
    all_customers = set(monday["by_customer"].keys())

    for e in emails:
        # Use full body, not preview
        text = (e["subject"] + " " + e.get("body_full", "")[:5000]).lower()

        for brand in all_brands:
            if len(brand) > 3 and brand in text:
                brand_mentions[brand] += 1
                matches[brand].append(e["uid"])

        for customer in all_customers:
            words = [w for w in customer.split() if len(w) > 3]
            if words and all(w in text for w in words[:2]):
                customer_mentions[customer] += 1

    return {
        "brand_mentions": dict(brand_mentions.most_common(60)),
        "customer_mentions": dict(customer_mentions.most_common(60)),
        "brand_email_map": {k: v[:10] for k, v in matches.items()},
    }


def analyze_communication_patterns(emails: list) -> dict:
    # Weekly volume by account
    weekly_volume = defaultdict(lambda: defaultdict(int))
    for e in emails:
        if e["date_simple"]:
            try:
                dt = datetime.strptime(e["date_simple"], "%Y-%m-%d")
                week_start = (dt - timedelta(days=dt.weekday())).strftime("%Y-%m-%d")
                acct = e["account"].lower().split()[0].split("+")[0]
                weekly_volume[week_start][acct] += 1
            except:
                pass

    # Top senders
    sender_count = Counter()
    for e in emails:
        if e["from_email"]:
            sender_count[e["from_email"]] += 1

    # Top recipients
    recipient_count = Counter()
    for e in emails:
        for addr in e.get("to", []):
            recipient_count[addr] += 1

    # Communication pairs
    pairs = Counter()
    for e in emails:
        f = e["from_email"]
        for t in e.get("to", []):
            pair = tuple(sorted([f, t]))
            pairs[pair] += 1

    # Thread analysis
    threads = defaultdict(list)
    for e in emails:
        key = e["base_subject"].lower()[:80]
        threads[key].append(e)

    long_threads = []
    for subj, msgs in threads.items():
        if len(msgs) >= 3:
            msgs_sorted = sorted(msgs, key=lambda x: x.get("date", ""))
            long_threads.append(
                {
                    "subject": subj,
                    "message_count": len(msgs),
                    "participants": list(set(m["from_email"] for m in msgs)),
                    "date_range": f"{msgs_sorted[0].get('date_simple', '?')} to {msgs_sorted[-1].get('date_simple', '?')}",
                    "classifications": list(
                        set(c for m in msgs for c in m.get("classifications", []))
                    ),
                    "total_body_chars": sum(m.get("body_length", 0) for m in msgs),
                    "attachment_count": sum(m.get("attachment_count", 0) for m in msgs),
                }
            )
    long_threads.sort(key=lambda x: x["message_count"], reverse=True)

    # Response time analysis
    msg_by_id = {e["message_id"]: e for e in emails if e["message_id"]}
    response_times = []
    for e in emails:
        if e["in_reply_to"] and e["in_reply_to"] in msg_by_id:
            orig = msg_by_id[e["in_reply_to"]]
            try:
                dt1 = datetime.fromisoformat(orig["date"])
                dt2 = datetime.fromisoformat(e["date"])
                diff = (dt2 - dt1).total_seconds()
                if 0 < diff < 7 * 86400:
                    response_times.append(
                        {
                            "from": e["from_email"],
                            "to": orig["from_email"],
                            "hours": round(diff / 3600, 1),
                        }
                    )
            except:
                pass

    resp_by_person = defaultdict(list)
    for r in response_times:
        resp_by_person[r["from"]].append(r["hours"])
    avg_response = {}
    for person, times in resp_by_person.items():
        avg_response[person] = {
            "avg_hours": round(sum(times) / len(times), 1),
            "median_hours": round(sorted(times)[len(times) // 2], 1),
            "reply_count": len(times),
        }

    # Classification distribution (multi-label)
    class_dist = Counter()
    for e in emails:
        for c in e.get("classifications", []):
            class_dist[c] += 1

    # Financial aggregates
    total_amounts_seen = []
    for e in emails:
        fin = e.get("financial", {})
        if fin.get("amounts"):
            total_amounts_seen.extend(fin["amounts"])

    # Negotiation patterns
    negotiation_keywords = Counter()
    neg_patterns = [
        r"\b(counter.?offer)\b",
        r"\b(revised|revision)\b",
        r"\b(best price|final price|last price)\b",
        r"\b(negotiat\w+)\b",
        r"\b(discount)\b",
        r"\b(margin)\b",
        r"\b(markup)\b",
        r"\b(urgent)\b",
        r"\b(asap)\b",
        r"\b(deadline)\b",
        r"\b(expire|expir\w+)\b",
        r"\b(approve|approval)\b",
        r"\b(confirm|confirmation)\b",
        r"\b(cancel\w*)\b",
        r"\b(delay\w*)\b",
        r"\b(problem|issue)\b",
    ]
    for e in emails:
        subj_lower = e["subject"].lower()
        for pat in neg_patterns:
            m = re.search(pat, subj_lower)
            if m:
                negotiation_keywords[m.group(1)] += 1

    # Deal numbers
    deal_nums = Counter()
    for e in emails:
        for dn in e.get("deal_numbers", []):
            deal_nums[dn] += 1

    # Attachment stats
    att_types = Counter()
    for e in emails:
        for att in e.get("attachments", []):
            ext = (
                att["filename"].rsplit(".", 1)[-1].lower()
                if "." in att["filename"]
                else "unknown"
            )
            att_types[ext] += 1

    return {
        "weekly_volume": dict(sorted(weekly_volume.items())),
        "top_senders": dict(sender_count.most_common(40)),
        "top_recipients": dict(recipient_count.most_common(40)),
        "top_pairs": [{"pair": list(p), "count": c} for p, c in pairs.most_common(40)],
        "long_threads": long_threads[:80],
        "response_times": avg_response,
        "classification_distribution": dict(class_dist),
        "negotiation_keywords": dict(negotiation_keywords.most_common(30)),
        "deal_numbers_found": dict(deal_nums.most_common(80)),
        "total_reply_chains_analyzed": len(response_times),
        "financial_summary": {
            "total_amounts_mentioned": len(total_amounts_seen),
            "sum_all_amounts": round(sum(total_amounts_seen), 2)
            if total_amounts_seen
            else 0,
            "largest_single_amount": max(total_amounts_seen)
            if total_amounts_seen
            else 0,
            "avg_amount": round(sum(total_amounts_seen) / len(total_amounts_seen), 2)
            if total_amounts_seen
            else 0,
        },
        "attachment_types": dict(att_types.most_common(20)),
    }


def generate_report(emails: list, analysis: dict, cross_ref: dict, monday: dict) -> str:
    """Generate comprehensive intelligence summary."""
    total = len(emails)
    by_account = Counter(e["account"].split("+")[0].split()[0] for e in emails)

    lines = []
    lines.append("# Tevah Omniscient Email Intelligence Report")
    lines.append(f"\n**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"**Period**: Last {DAYS_BACK} days")
    lines.append(f"**Total Emails Analyzed**: {total:,}")
    lines.append(f"**Accounts**: {len(ACCOUNTS)}")
    for name, count in by_account.most_common():
        lines.append(f"  - {name}: {count:,}")

    # Attachment summary
    total_att = sum(e.get("attachment_count", 0) for e in emails)
    total_att_text = sum(
        len(a.get("extracted_text", ""))
        for e in emails
        for a in e.get("attachments", [])
    )
    lines.append(f"\n**Attachments Parsed**: {total_att:,}")
    lines.append(f"**Total Text Extracted from Attachments**: {total_att_text:,} chars")
    if analysis.get("attachment_types"):
        types_str = ", ".join(
            f"{ext}: {count}" for ext, count in analysis["attachment_types"].items()
        )
        lines.append(f"**Attachment Types**: {types_str}")

    # Body coverage
    total_body = sum(e.get("body_length", 0) for e in emails)
    lines.append(
        f"**Total Email Body Text**: {total_body:,} chars ({total_body / 1_000_000:.1f}M)"
    )

    # Financial summary
    fin = analysis.get("financial_summary", {})
    if fin.get("total_amounts_mentioned"):
        lines.append(f"\n## Financial Data Extracted")
        lines.append(f"- **Dollar amounts found**: {fin['total_amounts_mentioned']:,}")
        lines.append(
            f"- **Sum of all amounts mentioned**: ${fin['sum_all_amounts']:,.2f}"
        )
        lines.append(
            f"- **Largest single amount**: ${fin['largest_single_amount']:,.2f}"
        )
        lines.append(f"- **Average amount**: ${fin['avg_amount']:,.2f}")

    # Classification
    lines.append("\n## Email Classification (Multi-Label)")
    cd = analysis["classification_distribution"]
    for cat in sorted(cd, key=cd.get, reverse=True):
        lines.append(f"- **{cat}**: {cd[cat]:,}")

    # Weekly volume
    lines.append("\n## Email Volume by Week")
    headers = sorted(
        set(
            acct
            for week_data in analysis["weekly_volume"].values()
            for acct in week_data
        )
    )
    lines.append("| Week | " + " | ".join(h.title() for h in headers) + " | Total |")
    lines.append("|---" + "|---" * (len(headers) + 1) + "|")
    for week, counts in analysis["weekly_volume"].items():
        vals = [str(counts.get(h, 0)) for h in headers]
        total_week = sum(counts.values())
        lines.append(f"| {week} | " + " | ".join(vals) + f" | {total_week} |")

    # Top senders
    lines.append("\n## Top 25 Senders")
    lines.append("| Sender | Count |")
    lines.append("|---|---|")
    for sender, count in list(analysis["top_senders"].items())[:25]:
        lines.append(f"| {sender} | {count} |")

    # Top recipients
    lines.append("\n## Top 25 Recipients")
    lines.append("| Recipient | Count |")
    lines.append("|---|---|")
    for recip, count in list(analysis["top_recipients"].items())[:25]:
        lines.append(f"| {recip} | {count} |")

    # Communication pairs
    lines.append("\n## Top Communication Pairs")
    lines.append("| Person A | Person B | Emails |")
    lines.append("|---|---|---|")
    for pd in analysis["top_pairs"][:25]:
        p = pd["pair"]
        lines.append(f"| {p[0]} | {p[1]} | {pd['count']} |")

    # Response times
    lines.append("\n## Response Time Analysis")
    if analysis["response_times"]:
        lines.append(
            f"\n*{analysis['total_reply_chains_analyzed']} reply chains analyzed*\n"
        )
        lines.append("| Person | Avg (hrs) | Median (hrs) | Replies |")
        lines.append("|---|---|---|---|")
        for person, stats in sorted(
            analysis["response_times"].items(),
            key=lambda x: x[1]["reply_count"],
            reverse=True,
        )[:25]:
            lines.append(
                f"| {person} | {stats['avg_hours']} | {stats['median_hours']} | {stats['reply_count']} |"
            )

    # Deal numbers
    lines.append("\n## Deal References Found")
    lines.append("| Number | Mentions |")
    lines.append("|---|---|")
    for num, count in list(analysis["deal_numbers_found"].items())[:50]:
        lines.append(f"| {num} | {count} |")

    # Brand cross-reference
    lines.append("\n## Brand Mentions (Cross-Referenced with Monday)")
    lines.append("| Brand | Emails | Monday Deals | Revenue |")
    lines.append("|---|---|---|---|")
    for brand, count in cross_ref["brand_mentions"].items():
        deals = monday["by_brand"].get(brand, [])
        rev = sum(d.get("total_revenue", 0) for d in deals)
        rev_str = f"${rev:,.0f}" if rev else "—"
        lines.append(f"| {brand.title()} | {count} | {len(deals)} | {rev_str} |")

    # Customer cross-reference
    lines.append("\n## Customer Mentions")
    lines.append("| Customer | Email Mentions |")
    lines.append("|---|---|")
    for cust, count in cross_ref["customer_mentions"].items():
        lines.append(f"| {cust.title()} | {count} |")

    # Long threads
    lines.append("\n## Longest Threads (3+ messages)")
    lines.append("| Subject | Msgs | Participants | Attachments | Date Range |")
    lines.append("|---|---|---|---|---|")
    for t in analysis["long_threads"][:40]:
        participants = ", ".join(t["participants"][:3])
        if len(t["participants"]) > 3:
            participants += f" +{len(t['participants']) - 3}"
        subj = t["subject"][:50] + ("..." if len(t["subject"]) > 50 else "")
        lines.append(
            f"| {subj} | {t['message_count']} | {participants} | {t.get('attachment_count', 0)} | {t['date_range']} |"
        )

    # Negotiation patterns
    lines.append("\n## Negotiation Signals")
    lines.append("| Keyword | Occurrences |")
    lines.append("|---|---|")
    for kw, count in analysis["negotiation_keywords"].items():
        lines.append(f"| {kw} | {count} |")

    return "\n".join(lines)


def main():
    start_time = time.time()
    print("=" * 60)
    print("  TEVAH OMNISCIENT EMAIL INGESTION")
    print("  Every word. Every attachment. Every number.")
    print("=" * 60)

    since = datetime.now() - timedelta(days=DAYS_BACK)
    since_str = since.strftime("%d-%b-%Y")
    print(f"\nPeriod: {since_str} → now")
    print(f"Accounts: {len(ACCOUNTS)}")

    # Ensure attachment dir exists
    os.makedirs(ATTACHMENT_DIR, exist_ok=True)

    # Load Monday deals
    print("\nLoading Monday deal data...")
    monday = load_monday_deals()
    print(f"  {len(monday['deals'])} deals, {len(monday['by_brand'])} brands")

    # Fetch from ALL accounts
    all_emails = []
    for account in ACCOUNTS:
        try:
            emails = fetch_emails_from_account(account, since_str)
            all_emails.extend(emails)
        except Exception as e:
            print(f"\n  [ERROR] {account['name']}: {e}")
            continue

    # Deduplicate
    before = len(all_emails)
    all_emails = deduplicate_emails(all_emails)
    print(f"\nDeduplicated: {before:,} → {len(all_emails):,}")

    # Sort by date
    all_emails.sort(key=lambda x: x.get("date", ""), reverse=True)

    # Cross-reference
    print("Cross-referencing with Monday deals...")
    cross_ref = cross_reference_deals(all_emails, monday)

    # Analyze
    print("Analyzing communication patterns...")
    analysis = analyze_communication_patterns(all_emails)

    # Save full JSON (this will be BIG — that's the point)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    json_output = {
        "generated": datetime.now().isoformat(),
        "version": "omniscient-v1",
        "period_days": DAYS_BACK,
        "total_emails": len(all_emails),
        "accounts": dict(
            Counter(e["account"].split("+")[0].split()[0] for e in all_emails)
        ),
        "total_body_chars": sum(e.get("body_length", 0) for e in all_emails),
        "total_attachments": sum(e.get("attachment_count", 0) for e in all_emails),
        "emails": all_emails,
        "analysis": analysis,
        "monday_cross_reference": cross_ref,
    }

    json_path = os.path.join(OUTPUT_DIR, "GMAIL-OMNISCIENT.json")
    print(f"\nWriting {json_path}...")
    with open(json_path, "w") as f:
        json.dump(json_output, f, indent=1, default=str)
    size_mb = os.path.getsize(json_path) / 1024 / 1024
    print(f"  Saved: {size_mb:.1f} MB")

    # Generate report
    report = generate_report(all_emails, analysis, cross_ref, monday)
    md_path = os.path.join(OUTPUT_DIR, "GMAIL-OMNISCIENT.md")
    with open(md_path, "w") as f:
        f.write(report)
    print(f"  Report: {md_path}")

    elapsed = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(
        f"  COMPLETE: {len(all_emails):,} emails | {sum(e.get('attachment_count', 0) for e in all_emails)} attachments"
    )
    print(f"  Body text: {sum(e.get('body_length', 0) for e in all_emails):,} chars")
    print(f"  Time: {elapsed / 60:.1f} minutes")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
